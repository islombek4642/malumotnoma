import sqlite3
import os

def create_database():
    # Papka mavjudligini tekshirish va yaratish
    if not os.path.exists("database"):
        os.makedirs("database")

    # Ma'lumotlar bazasini yaratish
    conn = sqlite3.connect("database/database.db")
    cursor = conn.cursor()

    # Jadvalni yaratish
    cursor.execute('''CREATE TABLE IF NOT EXISTS students (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        university TEXT,
        faculty TEXT,
        direction TEXT,
        course INTEGER,
        group_number TEXT,
        full_name TEXT,
        birth_date TEXT,
        passport_id TEXT,
        jshshir TEXT,
        status TEXT,
        education_type TEXT,
        education_form TEXT,
        phone TEXT,
        issued_date TEXT,
        valid_until TEXT
    )''')

    conn.commit()
    conn.close()

def get_student_by_id(identifier):
    conn = sqlite3.connect("database/database.db")
    cursor = conn.cursor()

    # Talabani passport_id yoki jshshir bo'yicha qidirish
    cursor.execute("SELECT * FROM students WHERE passport_id = ? OR jshshir = ?", (identifier, identifier))
    student = cursor.fetchone()

    conn.close()
    return student

def get_student_by_id_and_course(identifier, course):
    """Talabani pasport ID/JSHSHIR va kurs bo'yicha qidirish"""
    conn = sqlite3.connect("database/database.db")
    cursor = conn.cursor()

    # Talabani passport_id yoki jshshir va kurs bo'yicha qidirish
    cursor.execute("SELECT * FROM students WHERE (passport_id = ? OR jshshir = ?) AND course = ?",
                  (identifier, identifier, course))
    student = cursor.fetchone()

    conn.close()
    return student

def add_student(university, faculty, direction, course, group_number, full_name,
                birth_date, passport_id, jshshir, status, education_type,
                education_form, phone, issued_date, valid_until):
    try:
        conn = sqlite3.connect("database/database.db")
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO students (
                university, faculty, direction, course, group_number, full_name,
                birth_date, passport_id, jshshir, status, education_type,
                education_form, phone, issued_date, valid_until
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (university, faculty, direction, course, group_number, full_name,
              birth_date, passport_id, jshshir, status, education_type,
              education_form, phone, issued_date, valid_until))

        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Ma'lumotlarni qo'shishda xatolik: {e}")
        return False

def get_all_students():
    conn = sqlite3.connect("database/database.db")
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM students")
    students = cursor.fetchall()

    conn.close()
    return students

def import_students_from_excel(file_path):
    """Excel fayldan talabalar ma'lumotlarini import qilish"""
    try:
        import openpyxl

        conn = sqlite3.connect("database/database.db")
        cursor = conn.cursor()

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        success_count = 0
        error_count = 0

        # Xatoliklar statistikasi
        empty_cells_count = 0
        missing_columns_count = 0
        extra_columns_count = 0

        # Kerakli ustunlar soni
        required_columns = 15

        # Birinchi qatorni o'tkazib yuborish (sarlavhalar)
        rows = list(sheet.rows)
        if len(rows) <= 1:
            return 0, 0, "Faylda faqat sarlavhalar mavjud yoki fayl bo'sh"

        for row_index, row in enumerate(rows[1:], 2):  # 2-qatordan boshlab (1-qator sarlavhalar)
            try:
                # Qatordan ma'lumotlarni olish
                cells = [cell.value for cell in row]

                # Ustunlar sonini tekshirish
                if len(cells) < required_columns:
                    missing_columns_count += 1
                    error_count += 1
                    print(f"Qator {row_index}: Ustunlar yetishmaydi. Kerak: {required_columns}, Mavjud: {len(cells)}")
                    continue
                elif len(cells) > required_columns:
                    extra_columns_count += 1
                    # Ortiqcha ustunlarni kesib tashlash
                    cells = cells[:required_columns]
                    print(f"Qator {row_index}: Ortiqcha ustunlar mavjud. Faqat birinchi {required_columns} ta ustun olinadi.")

                # Bo'sh ustunlarni tekshirish
                if None in cells or "" in cells:
                    empty_cells_count += 1
                    error_count += 1
                    empty_indices = [i+1 for i, cell in enumerate(cells) if cell is None or cell == ""]
                    print(f"Qator {row_index}: Bo'sh ustunlar mavjud: {empty_indices}")
                    continue

                university, faculty, direction, course, group_number, full_name, \
                birth_date, passport_id, jshshir, status, education_type, \
                education_form, phone, issued_date, valid_until = cells

                # Kursni raqamga aylantirish
                try:
                    course = int(course)
                except (ValueError, TypeError):
                    error_count += 1
                    print(f"Qator {row_index}: Kurs raqam bo'lishi kerak")
                    continue

                # Ma'lumotlarni bazaga qo'shish
                cursor.execute('''
                    INSERT INTO students (
                        university, faculty, direction, course, group_number, full_name,
                        birth_date, passport_id, jshshir, status, education_type,
                        education_form, phone, issued_date, valid_until
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (university, faculty, direction, course, group_number, full_name,
                      birth_date, passport_id, jshshir, status, education_type,
                      education_form, phone, issued_date, valid_until))

                success_count += 1
            except Exception as e:
                print(f"Qator {row_index} import qilishda xatolik: {e}")
                error_count += 1

        conn.commit()
        conn.close()

        # Xatoliklar haqida batafsil ma'lumot
        error_details = ""
        if empty_cells_count > 0:
            error_details += f"\n- {empty_cells_count} ta qatorda bo'sh ustunlar mavjud"
        if missing_columns_count > 0:
            error_details += f"\n- {missing_columns_count} ta qatorda ustunlar yetishmaydi"
        if extra_columns_count > 0:
            error_details += f"\n- {extra_columns_count} ta qatorda ortiqcha ustunlar mavjud"

        return success_count, error_count, error_details
    except Exception as e:
        print(f"Excel import qilishda xatolik: {e}")
        return 0, 0, f"Xatolik: {str(e)}"

def create_sample_excel():
    """Namuna Excel fayl yaratish"""
    try:
        import openpyxl
        from openpyxl.styles import Font

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # O'zbekcha sarlavhalar
        headers = [
            "Universitet", "Fakultet", "Yo'nalish", "Kurs", "Guruh", "F.I.O",
            "Tug'ilgan sana", "Pasport ID", "JSHSHIR", "Talabalik holati", "O'qish turi",
            "O'qish shakli", "Telefon", "Berilgan sana", "Amal qilish muddati"
        ]

        # Sarlavhalarni yozish
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        # Namuna ma'lumotlar
        sample_data = [
            "Toshkent Davlat Universiteti", "Kompyuter fakulteti", "Dasturlash", 2, "201-guruh",
            "Aliyev Ali Aliyevich", "01.01.2000", "AA1234567", "12345678901234", "aktiv",
            "davlat granti", "kunduzgi", "+998901234567", "01.09.2023", "01.07.2024"
        ]

        # Namuna ma'lumotlarni yozish
        for col, value in enumerate(sample_data, 1):
            sheet.cell(row=2, column=col).value = value

        # Faylni saqlash
        file_path = "namuna_talabalar.xlsx"
        workbook.save(file_path)
        return file_path
    except Exception as e:
        print(f"Namuna fayl yaratishda xatolik: {e}")
        return None

def delete_student_by_id(student_id):
    """Talabani ID bo'yicha o'chirish"""
    try:
        conn = sqlite3.connect("database/database.db")
        cursor = conn.cursor()

        cursor.execute("DELETE FROM students WHERE id = ?", (student_id,))

        deleted = cursor.rowcount > 0
        conn.commit()
        conn.close()

        return deleted
    except Exception as e:
        print(f"Talabani o'chirishda xatolik: {e}")
        return False

def delete_all_students():
    """Barcha talabalarni o'chirish"""
    try:
        conn = sqlite3.connect("database/database.db")
        cursor = conn.cursor()

        cursor.execute("DELETE FROM students")

        deleted_count = cursor.rowcount
        conn.commit()
        conn.close()

        return deleted_count
    except Exception as e:
        print(f"Barcha talabalarni o'chirishda xatolik: {e}")
        return 0

def delete_student_by_passport(passport_id):
    """Talabani pasport ID bo'yicha o'chirish"""
    try:
        conn = sqlite3.connect("database/database.db")
        cursor = conn.cursor()

        # Avval talabani topish
        cursor.execute("SELECT id, full_name FROM students WHERE passport_id = ?", (passport_id,))
        student = cursor.fetchone()

        if not student:
            conn.close()
            return None

        # Talabani o'chirish
        cursor.execute("DELETE FROM students WHERE passport_id = ?", (passport_id,))

        conn.commit()
        conn.close()

        return student
    except Exception as e:
        print(f"Talabani pasport ID bo'yicha o'chirishda xatolik: {e}")
        return None

# Ma'lumotlar bazasini yaratish
create_database()